import json
from openpyxl import Workbook
from openpyxl import load_workbook
import re
import requests
from statistics import mean, median, mode
import matplotlib.pyplot as plt
import random

CHISTES = {
    "error" : "",
    "category" : 0,
    "type" : 0,
    "setup" : "",
    "delivery": "",
    "id": "",
    "safe": "",
    "lang": ""
}



DatosProcesados = {
    "chistes_dos_partes": "",
    "prom_chistes_dos_lineas": "",
    "prom_exito_ejecucion": ""
    }



ListaDatosProcesados = []
ListaDatosProcesados.append(DatosProcesados)

      
listaChistes = []
listaDatos = []

with open("Avance2Archivo.txt", "r") as file:
    lineas = file.readlines()#funcion que lee cada linea y devuelve lista de str
    
    
    lineas_chistes = lineas[:-1]   #todas las lineas del archivo menos la última (o sea los chistes)
    linea_datos = lineas[-1]       #ultima linea (las consultas que calculamos)

    for line in lineas_chistes:
        line = line.replace('\n', '')
        info_chiste = line.split(',')

        
        CHISTES["error"] = info_chiste[0]
        CHISTES["category"] = info_chiste[1]
        CHISTES["type"] = info_chiste[2]
        CHISTES["setup"] = info_chiste[3]
        CHISTES["delivery"] = info_chiste[4]
        CHISTES["safe"] = info_chiste[5]
        CHISTES["id"] = int(info_chiste[6])
        CHISTES["lang"] = info_chiste[7]
        copia = CHISTES.copy()
        listaChistes.append(copia)

    #ultima linea del archivo que contiene las consultas
    linea_datos = linea_datos.replace('\n', '')
    info_datos = linea_datos.split(',')

    
    DatosProcesados["chistes_dos_partes"] = int(info_datos[0])
    DatosProcesados["prom_chistes_dos_lineas"] =float(info_datos[1])
    DatosProcesados["prom_exito_ejecucion"] = float(info_datos[2])
    copia2 = DatosProcesados.copy()
    listaDatos.append(copia2)
    
#print("------------")
print("lista de chistes leída desde el archivo", listaChistes)
print("lista de consultas leída desde el archivo", listaDatos)



"""EXPRESIONES REGULARES"""

chistes2 = listaChistes.copy()
texto = str(chistes2)

DetectorId = re.compile(r"\d+")
IdValido = DetectorId.findall(texto)
#print(IdValido)
if len(IdValido) == 0:
    print("Hubo un fallo en los ID")
else:
    print("No hubieron errores en los IDS, todos son numeros según las expresiones regulares")

#reorganizacion de los datos categoria e id
#calculo de media y mediana de los id

listaIDS= []

for i in range(len(listaChistes)):
        listaIDS.append(listaChistes[i]["id"])
print(listaIDS)

mean = mean(listaIDS)
median = median(listaIDS)
print("Media de los IDs:  ", mean)

print("Mediana de los IDs:  ", median)

#calculo de la moda de las categorías
listaCATEGORIAS = []
for i in range(len(listaChistes)):
        listaCATEGORIAS.append(listaChistes[i]["category"])
print(listaCATEGORIAS)
mode = mode(listaCATEGORIAS)
print("La moda de las categorías", mode)


#GUARDAR EN EXCEL listaChistes y listaDatos
libro = Workbook()         
hoja1 = libro.active        
nombre = "hoja1"               
hoja1.title = nombre
hoja2 = libro.create_sheet("hoja2")


try:
    libro = load_workbook("PIA.xlsx") #este tramo lo tomamos de un script de la clase
except FileNotFoundError:
    print("el archivo no existe")

hoja1["A1"] = "error"
hoja1["B1"] = "category"
hoja1["C1"] = "type"
hoja1["D1"] = "setup"
hoja1["E1"] = "delivery"
hoja1["F1"] = "id"
hoja1["G1"] = "safe"
hoja1["H1"] = "lang"

count = 2
for i in range(len(listaChistes)): 
    hoja1.cell(count,1,listaChistes[i]['error'])
    hoja1.cell(count,2,listaChistes[i]['category'])
    hoja1.cell(count,3,listaChistes[i]['type'])
    hoja1.cell(count,4,listaChistes[i]['setup'])
    hoja1.cell(count,5,listaChistes[i]['delivery'])
    hoja1.cell(count,6,listaChistes[i]['id'])
    hoja1.cell(count,7,listaChistes[i]['safe'])
    hoja1.cell(count,8,listaChistes[i]['lang'])
    count += 1


hoja2["A1"] = "chistes_dos_partes"
hoja2["B1"] = "prom_chistes_dos_lineas"
hoja2["C1"] = "prom_exito_ejecucion"

count = 2
for i in range(len(listaDatos)): 
    hoja2.cell(count,1,listaDatos[i]['chistes_dos_partes'])
    hoja2.cell(count,2,listaDatos[i]['prom_chistes_dos_lineas'])
    hoja2.cell(count,3,listaDatos[i]['prom_exito_ejecucion'])
    count += 1

libro.save("PIA.xlsx")

#Graficas
"""grafica1"""
y = []
x = list(range(len(listaIDS)))
num = len(listaIDS)
for i in range(num):
    y.append(random.randint(1,1000))
plt.plot(x, y, marker='o')
plt.title('posicion (ids) completamente aleatorio')
plt.xlabel('posición id')
plt.ylabel('numeros aleatorios')
plt.show()


"""grafica2"""
x= listaIDS
y = list(range(len(listaIDS)))
plt.scatter(x, y)
plt.title('dispersión de IDS')
plt.xlabel('posicion del id procesado')
plt.ylabel('numero')
plt.show()


"""grafica3"""
x = list(range(len(listaIDS)))
y = listaIDS
plt.bar(x, y)
plt.title('numeros de los ids graficamente')
plt.xlabel('id posicion en lista')
plt.ylabel('numero')
plt.show()


"""grafica4"""
categorias_posibles = ['Pun', 'Programming', 'Spooky', 'Misc', 'Christmas']
frecuencias = []
for categoria in categorias_posibles:
    frecuencias.append(listaCATEGORIAS.count(categoria)) #cuenta cuántas veces aparece un
                                                          #valor especifico en una lista
plt.pie(frecuencias, labels=categorias_posibles)
plt.title('moda de categorías')
plt.show()

